import io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

AZUL       = "0E4F8A"
AZUL_CLARO = "D6E4F0"
GRIS       = "F2F2F2"
BLANCO     = "FFFFFF"

_thin  = Side(style="thin",   color="AAAAAA")
_med   = Side(style="medium", color=AZUL)
_borde = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_borde_med = Border(left=_med, right=_med, top=_med, bottom=_med)


def _hdr(ws, row, col, text, span=1, bg=AZUL, color=BLANCO, sz=12, bold=True, h=22):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Arial", bold=bold, size=sz, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _borde_med
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    ws.row_dimensions[row].height = h


def _lbl(ws, row, col, text, span=1, bg=AZUL_CLARO, bold=True, h=15):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Arial", bold=bold, size=9, color=AZUL)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = _borde
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    ws.row_dimensions[row].height = h


def _val(ws, row, col, text="", span=1, h=15, bg=BLANCO, bold=False, sz=9):
    cell = ws.cell(row=row, column=col, value=text or "")
    cell.font = Font(name="Arial", size=sz, bold=bold)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = _borde
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    ws.row_dimensions[row].height = h


def _blank(ws, row, col, span=1, h=30, bg=BLANCO):
    _val(ws, row, col, "", span=span, h=h, bg=bg)


def generar_informe_final(data) -> bytes:
    dat  = data.datos if hasattr(data, 'datos') else type('', (), data.get('datos', {}))()
    obj  = data.objetivos if hasattr(data, 'objetivos') else type('', (), data.get('objetivos', {}))()

    nombre_curso = getattr(dat, 'nombreCurso', '') or ""
    instructor   = getattr(dat, 'instructor',  '') or ""
    lugar        = getattr(dat, 'lugar',       '') or ""
    fecha        = getattr(dat, 'fecha',       '') or ""
    duracion     = getattr(dat, 'duracion',    '') or ""
    participantes = int(getattr(dat, 'participantes', None) or 10)

    obj_general   = getattr(obj, 'general',    '') or ""
    obj_cognitiva = getattr(obj, 'cognitiva',  '') or ""
    obj_psico     = getattr(obj, 'psicomotriz','') or ""
    obj_afectiva  = getattr(obj, 'afectiva',   '') or ""
    obj_partic = f"1. (Cognitivo) {obj_cognitiva}\n2. (Psicomotriz) {obj_psico}\n3. (Afectivo) {obj_afectiva}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Informe Final"

    # Anchos de columna (A–H)
    for col, w in enumerate([28, 18, 6, 18, 6, 14, 6, 14], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    r = 1
    # ── Título ────────────────────────────────────────────────────────────────
    _hdr(ws, r, 1, "INFORME FINAL DEL CURSO", span=8, sz=14, h=28); r += 1

    # ── Datos generales ───────────────────────────────────────────────────────
    _lbl(ws, r, 1, "Nombre del curso/sesión:")
    _val(ws, r, 2, nombre_curso, span=7); r += 1

    _lbl(ws, r, 1, "Instructor:")
    _val(ws, r, 2, instructor, span=3)
    _lbl(ws, r, 5, "Lugar:")
    _val(ws, r, 6, lugar, span=3); r += 1

    _lbl(ws, r, 1, "Fecha:")
    _val(ws, r, 2, fecha)
    _lbl(ws, r, 3, "Duración:")
    _val(ws, r, 4, f"{duracion} min" if duracion else "")
    _lbl(ws, r, 5, "Horario:")
    _val(ws, r, 6, "")
    _lbl(ws, r, 7, "a:")
    _val(ws, r, 8, ""); r += 1

    # ── Comentarios del instructor ────────────────────────────────────────────
    _hdr(ws, r, 1, "Comentarios del Instructor acerca del proceso de aprendizaje del curso",
         span=8, bg=AZUL, sz=10, h=18); r += 1
    _blank(ws, r, 1, span=8, h=50); r += 1

    # ── Cumplimiento de objetivos ─────────────────────────────────────────────
    _hdr(ws, r, 1,
         "Nivel de cumplimiento de los objetivos/resultados de aprendizaje y de las expectativas del curso",
         span=8, bg=AZUL, sz=10, h=18); r += 1

    _lbl(ws, r, 1, "Objetivo General", span=4)
    _lbl(ws, r, 5, "¿Al finalizar el curso se revisaron los objetivos?", span=4); r += 1
    _val(ws, r, 1, obj_general, span=4, h=40)
    _blank(ws, r, 5, span=4, h=40); r += 1

    _lbl(ws, r, 1, "Objetivos Particulares", span=4)
    _lbl(ws, r, 5, "¿Al finalizar el curso se revisaron los objetivos?", span=4); r += 1
    _val(ws, r, 1, obj_partic, span=4, h=55)
    _blank(ws, r, 5, span=4, h=55); r += 1

    _lbl(ws, r, 1, "Expectativas de los participantes", span=2)
    _lbl(ws, r, 3, "Descripción del nivel de cumplimiento de las expectativas", span=6); r += 1
    _blank(ws, r, 1, span=2, h=35)
    _blank(ws, r, 3, span=6, h=35); r += 1

    # ── Plan de seguimiento ───────────────────────────────────────────────────
    _hdr(ws, r, 1,
         "Plan de seguimiento a los participantes en la aplicación de lo aprendido",
         span=8, bg=AZUL, sz=10, h=18); r += 1
    _lbl(ws, r, 1, "Nombre del Participante", span=2)
    _lbl(ws, r, 3, "Calificación final", span=2)
    _lbl(ws, r, 5, "Sugerencia de aprendizaje", span=2)
    _lbl(ws, r, 7, "Seguimiento", span=2); r += 1
    for _ in range(participantes):
        _blank(ws, r, 1, span=2)
        _blank(ws, r, 3, span=2)
        _blank(ws, r, 5, span=2)
        _blank(ws, r, 7, span=2); r += 1

    # ── Contingencias ─────────────────────────────────────────────────────────
    _hdr(ws, r, 1, "Contingencias presentadas durante el curso/sesión",
         span=8, bg=AZUL, sz=10, h=18); r += 1
    _lbl(ws, r, 1, "¿Se presentó alguna contingencia?", span=2)
    _lbl(ws, r, 3, "SÍ", span=1)
    _blank(ws, r, 4, span=1)
    _lbl(ws, r, 5, "NO", span=1)
    _blank(ws, r, 6, span=3); r += 1
    _lbl(ws, r, 1, "¿Cómo se resolvieron?", span=2)
    _blank(ws, r, 3, span=6, h=35); r += 1
    _lbl(ws, r, 1, "Ajustes al plan de sesión implementados:", span=2)
    _blank(ws, r, 3, span=6, h=35); r += 1

    # ── Encuesta de satisfacción ──────────────────────────────────────────────
    _hdr(ws, r, 1,
         "Resultado de la encuesta de satisfacción / Resumen de recomendaciones para mejora del curso",
         span=8, bg=AZUL, sz=10, h=18); r += 1
    _lbl(ws, r, 1, "Apartado", span=3)
    _lbl(ws, r, 4, "Nivel de satisfacción", span=2)
    _lbl(ws, r, 6, "Sugerencias de mejora", span=3); r += 1
    for apt in ["Características del evento", "Contenidos del curso", "Instalaciones",
                "Material didáctico", "Recursos empleados", "Desempeño del Instructor"]:
        _val(ws, r, 1, apt, span=3)
        _blank(ws, r, 4, span=2)
        _blank(ws, r, 6, span=3); r += 1

    # ── Resultados de evaluación ──────────────────────────────────────────────
    _hdr(ws, r, 1,
         "Resultados de las evaluaciones de aprendizaje",
         span=8, bg=AZUL, sz=10, h=18); r += 1
    _lbl(ws, r, 1, "Nombre del participante", span=2)
    _lbl(ws, r, 3, "Diagnóstica", span=2)
    _lbl(ws, r, 5, "Formativa", span=2)
    _lbl(ws, r, 7, "Final (Sumativa)", span=1)
    _lbl(ws, r, 8, "Total", span=1); r += 1
    for _ in range(participantes):
        _blank(ws, r, 1, span=2)
        _blank(ws, r, 3, span=2)
        _blank(ws, r, 5, span=2)
        _blank(ws, r, 7, span=1)
        _blank(ws, r, 8, span=1); r += 1

    # ── Avances logrados ──────────────────────────────────────────────────────
    _lbl(ws, r, 1, "Avances logrados:", span=8); r += 1
    _blank(ws, r, 1, span=8, h=45); r += 1

    # ── Firma ─────────────────────────────────────────────────────────────────
    _lbl(ws, r, 1, "Nombre y firma del instructor:", span=3)
    _val(ws, r, 4, instructor, span=5, h=20); r += 1

    ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
